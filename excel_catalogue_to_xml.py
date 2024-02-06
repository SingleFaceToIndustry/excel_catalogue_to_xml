"""
 * Copyright (C) 2023 SFTI and Swedish Local Authorities and Regions (SALAR)
 *
 * Licensed under the Apache License, Version 2.0 (the "License");
 * you may not use this file except in compliance with the License.
 * You may obtain a copy of the License at
 *
 *         http://www.apache.org/licenses/LICENSE-2.0
 *
 * Unless required by applicable law or agreed to in writing, software
 * distributed under the License is distributed on an "AS IS" BASIS,
 * WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
 * See the License for the specific language governing permissions and
 * limitations under the License.
 */
"""
from helper_functions import *
import io
from openpyxl import load_workbook
import configparser
import xml.etree.ElementTree as el_tree
import warnings


def excel_to_xml(excel_file, max_line_items=None) -> str:
    '''
    Takes an excel spread sheet and transforms it into Peppol BIS Catalogue XML
    :param max_line_items: Maximum number of line items to process
    :param excel_file: file path to the file or byte-array containing the file
    :return: XML-string
    '''
    try:
        # Filter warnings from openpyxl
        warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

        ## Load Excel workbook, test if the input is a path or a not (byte-array)
        if isinstance(excel_file, str):
            wb = load_workbook(filename=excel_file)
        else:
            wb = load_workbook(filename=io.BytesIO(excel_file))
    except Exception as e:
        raise ValueError("Not a valid Excel file") from e


    # Load config parser with parameters for all business terms locations and codelists
    config = configparser.ConfigParser()
    config.read("ExcelCellLocations.cfg")

    # Verify the consistency of the template (all sheets in place, all columns in correct order and so forth)
    check_spreadsheet_consistency(wb)

    # Assign the main spreadsheets to variables
    sheet_header = wb["CatalogueHeader"]
    sheet_lines = wb["CatalogueLines"]

    # Create Codelists objects from the codelists in the spreadsheet
    country_codes = load_code_list(wb, cl_range(config, "LIST_COUNTRY_CODE"))
    price_type_codes = load_code_list(wb, cl_range(config, "LIST_PRICE_TYPE"))
    vat_codes = load_code_list(wb, cl_range(config, "LIST_VAT_CODE"))
    unit_codes = load_code_list(wb, cl_range(config, "LIST_UNIT_CODE"))
    item_classification_codes = load_code_list(wb, cl_range(config, "LIST_ITEM_CLASSIFICATION_CODE"))
    item_property_codes = load_code_list(wb, cl_range(config, "LIST_ITEM_PROPERTY_CODE"))
    item_attribute_codes = load_code_list(wb, cl_range(config, "LIST_ITEM_ATTRIBUTE_CODE"))
    item_measure_codes = load_code_list(wb, cl_range(config, "LIST_ITEM_MEASURE_CODE"))
    item_certificate_env_codes = load_code_list(wb, cl_range(config, "LIST_ITEM_CERTIFICATE_ENV_CODE"))
    item_certificate_nutr_codes = load_code_list(wb, cl_range(config, "LIST_ITEM_CERTIFICATE_NUTR_CODE"))
    item_availability_codes = load_code_list(wb, cl_range(config, "LIST_ITEM_AVAILABILITY_CODE"))

    # Create root element with namespaces
    root = el_tree.Element("Catalogue")
    root.set("xmlns", "urn:oasis:names:specification:ubl:schema:xsd:Catalogue-2")
    root.set("xmlns:cac", "urn:oasis:names:specification:ubl:schema:xsd:CommonAggregateComponents-2")
    root.set("xmlns:cbc", "urn:oasis:names:specification:ubl:schema:xsd:CommonBasicComponents-2")

    # Header Information
    add_element(el_tree, root, "cbc:CustomizationID", "urn:fdc:peppol.eu:poacc:trns:catalogue:3")
    add_element(el_tree, root, "cbc:ProfileID", "urn:fdc:peppol.eu:poacc:bis:catalogue_wo_response:3")
    add_element(el_tree, root, "cbc:ID", str(sheet_header[header_cell(config, "CATALOGUE_ID")].value))
    add_element(el_tree, root, "cbc:ActionCode", str(sheet_header[header_cell(config, "ACTIONCODE")].value))
    add_element(el_tree, root, "cbc:Name", str(sheet_header[header_cell(config, "CATALOGUE_NAME")].value))
    add_element(el_tree, root, "cbc:IssueDate", str(sheet_header[header_cell(config, "CATALOGUE_ISSUEDATE")].value).split(" ")[0])

    currency_id = str(sheet_header[header_cell(config, "CURRENCY_ID")].value)

    # Validity
    if not is_cell_empty(str(sheet_header[header_cell(config, "CATALOGUE_STARTDATE")].value)) or not is_cell_empty(str(sheet_header[header_cell(config, "CATALOGUE_ENDDATE")].value)):
        cac = el_tree.SubElement(root, "cac:ValidityPeriod")
        add_element(el_tree, cac, "cbc:StartDate", str(sheet_header[header_cell(config, "CATALOGUE_STARTDATE")].value).split(" ")[0])
        add_element(el_tree, cac, "cbc:EndDate", str(sheet_header[header_cell(config, "CATALOGUE_ENDDATE")].value).split(" ")[0])

    # Referenced contract
    if not is_cell_empty(str(sheet_header[header_cell(config, "REFERENCED_CONTRACT_ID")].value)):
        cac = el_tree.SubElement(root, "cac:ReferencedContract")
        add_element(el_tree, cac, "cbc:ID", str(sheet_header[header_cell(config, "REFERENCED_CONTRACT_ID")].value))

    # Previous catalogue (source catalogue)
    if not is_cell_empty(str(sheet_header[header_cell(config, "PREVIOUS_CATALOGUE_ID")].value)):
        cac = el_tree.SubElement(root, "cac:SourceCatalogueReference")
        add_element(el_tree, cac, "cbc:ID", str(sheet_header[header_cell(config, "PREVIOUS_CATALOGUE_ID")].value))

    # PROVIDER Supplier party - sender
    cac = el_tree.SubElement(root, "cac:ProviderParty")
    c = add_element(el_tree, cac, "cbc:EndpointID", str(sheet_header[header_cell(config, "PROVIDER_SUPPLIER_ENDPOINT_ID")].value))
    add_attribute(c, "schemeID", str(sheet_header[header_cell(config, "PROVIDER_SUPPLIER_ENDPOINT_ID_SCHEMEID")].value))
    cac1 = el_tree.SubElement(cac, "cac:PartyIdentification")
    c = add_element(el_tree, cac1, "cbc:ID", str(sheet_header[header_cell(config, "PROVIDER_SUPPLIER_PARTY_ID")].value))
    add_attribute(c, "schemeID", str(sheet_header[header_cell(config, "PROVIDER_SUPPLIER_PARTY_ID_SCHEMEID")].value))
    c = el_tree.SubElement(cac, "cac:PartyLegalEntity")
    add_element(el_tree, c, "cbc:RegistrationName", str(sheet_header[header_cell(config, "PROVIDER_SUPPLIER_NAME")].value))

    # RECEIVER Buyer party - receiver
    cac = el_tree.SubElement(root, "cac:ReceiverParty")
    c = add_element(el_tree, cac, "cbc:EndpointID", str(sheet_header[header_cell(config, "RECEIVER_BUYER_ENDPOINT_ID")].value))
    add_attribute(c, "schemeID", str(sheet_header[header_cell(config, "RECEIVER_BUYER_ENDPOINT_ID_SCHEMEID")].value))
    cac1 = el_tree.SubElement(cac, "cac:PartyIdentification")
    c = add_element(el_tree, cac1, "cbc:ID", str(sheet_header[header_cell(config, "RECEIVER_BUYER_PARTY_ID")].value))
    add_attribute(c, "schemeID", str(sheet_header[header_cell(config, "RECEIVER_BUYER_PARTY_ID_SCHEMEID")].value))
    c = el_tree.SubElement(cac, "cac:PartyLegalEntity")
    add_element(el_tree, c, "cbc:RegistrationName", str(sheet_header[header_cell(config, "RECEIVER_BUYER_NAME")].value))

    # Supplier party
    if not is_cell_empty(str(sheet_header[header_cell(config, "SUPPLIER_NAME")].value)):
        cac = el_tree.SubElement(root, "cac:SellerSupplierParty")
        cac = el_tree.SubElement(cac, "cac:Party")
        c = add_element(el_tree, cac, "cbc:EndpointID", str(sheet_header[header_cell(config, "SUPPLIER_ENDPOINT_ID")].value))
        add_attribute(c, "schemeID", str(sheet_header[header_cell(config, "SUPPLIER_ENDPOINT_ID_SCHEMEID")].value))
        cac1 = el_tree.SubElement(cac, "cac:PartyIdentification")
        c = add_element(el_tree, cac1, "cbc:ID", str(sheet_header[header_cell(config, "SUPPLIER_PARTY_ID")].value))
        add_attribute(c, "schemeID", str(sheet_header[header_cell(config, "SUPPLIER_PARTY_ID_SCHEMEID")].value))
        c = el_tree.SubElement(cac, "cac:PartyName")
        add_element(el_tree, c, "cbc:Name", str(sheet_header[header_cell(config, "SUPPLIER_NAME")].value))

    # Buyer party
    if not is_cell_empty(str(sheet_header[header_cell(config, "BUYER_NAME")].value)):
        cac = el_tree.SubElement(root, "cac:ContractorCustomerParty")
        cac = el_tree.SubElement(cac, "cac:Party")
        c = add_element(el_tree, cac, "cbc:EndpointID", str(sheet_header[header_cell(config, "BUYER_ENDPOINT_ID")].value))
        add_attribute(c, "schemeID", str(sheet_header[header_cell(config, "BUYER_ENDPOINT_ID_SCHEMEID")].value))
        cac1 = el_tree.SubElement(cac, "cac:PartyIdentification")
        c = add_element(el_tree, cac1, "cbc:ID", str(sheet_header[header_cell(config, "BUYER_PARTY_ID")].value))
        add_attribute(c, "schemeID", str(sheet_header[header_cell(config, "BUYER_PARTY_ID_SCHEMEID")].value))
        c = el_tree.SubElement(cac, "cac:PartyName")
        add_element(el_tree, c, "cbc:Name", str(sheet_header[header_cell(config, "BUYER_NAME")].value))

    processed_lines = 0
    # Loop through rows in the CatalogueLine sheet starting from row 3 according to the template
    for row in sheet_lines.iter_rows(min_row=3, max_row=sheet_lines.max_row):

        # if no line number, then assume an empty or incomplete row and exit the loop.
        if row[col_index(config, "LINE_ID")].value is None:
            break
        elif str(row[col_index(config, "LINE_ID")].value).lower == "x":
            # In case the line number cell is x, then skip the line and continue with next
            continue

        # In case the provided value for  max number of line items have been reached, then exit the loop
        if isinstance(max_line_items, int):
            processed_lines += 1
            if processed_lines > max_line_items:
                break

        cac_CatalogueLine = el_tree.SubElement(root, "cac:CatalogueLine")

        # Sub-elements under cac:CatalogueLine
        add_element(el_tree, cac_CatalogueLine, "cbc:ID", str(row[col_index(config, "LINE_ID")].value))
        add_element(el_tree, cac_CatalogueLine, "cbc:ActionCode", "Add")

        # if OrderableIndicator is empty in the spread sheet, then set value true
        add_element(el_tree, cac_CatalogueLine, "cbc:OrderableIndicator", "false" if str(row[col_index(config, "ORDERABLEINDICATOR")].value).lower() == "nej" else "true")
        add_element(el_tree, cac_CatalogueLine, "cbc:OrderableUnit", get_code(str(row[col_index(config, "BASEQUANTITY_CODE")].value), unit_codes))

        c = add_element(el_tree, cac_CatalogueLine, "cbc:ContentUnitQuantity", str(row[col_index(config, "CONTENTUNITQUANTITY")].value))
        add_attribute(c, "unitCode", get_code(str(row[col_index(config, "CONTENTUNITQUANTITY_CODE")].value), unit_codes))

        add_element(el_tree, cac_CatalogueLine, "cbc:OrderQuantityIncrementNumeric", str(row[col_index(config, "ORDERQUANTITYINCREMENTNUMERIC")].value))

        c = add_element(el_tree, cac_CatalogueLine, "cbc:MinimumOrderQuantity",
                        str(row[col_index(config, "MINIMUMORDERQUANTITY")].value))
        add_attribute(c, "unitCode", get_code(str(row[col_index(config, "ORDERABLEUNIT")].value), unit_codes))

        add_element(el_tree, cac_CatalogueLine, "cbc:PackLevelCode", str(row[col_index(config, "PACKLEVELCODE")].value))

        if not is_cell_empty(str(row[col_index(config, "LINE_VALIDITY_STARTDATE")].value)) or not is_cell_empty(str(row[col_index(config, "LINE_VALIDITY_ENDDATE")].value)):
            cac = el_tree.SubElement(cac_CatalogueLine, "cac:LineValidityPeriod")
            add_element(el_tree, cac, "cbc:StartDate", str(row[col_index(config, "LINE_VALIDITY_STARTDATE")].value).split(" ")[0])
            add_element(el_tree, cac, "cbc:EndDate", str(row[col_index(config, "LINE_VALIDITY_ENDDATE")].value).split(" ")[0])

        if not is_cell_empty(str(row[col_index(config, "ITEMCOM_PRICEAMOUNT")].value)) or not is_cell_empty(str(row[col_index(config, "ITEMCOM_QUANTITY")].value)):
            cac = el_tree.SubElement(cac_CatalogueLine, "cac:ItemComparison")
            c = add_element(el_tree, cac, "cbc:PriceAmount", str(row[col_index(config, "ITEMCOM_PRICEAMOUNT")].value))
            add_attribute(c, "currencyID", currency_id)
            c = add_element(el_tree, cac, "cbc:Quantity", str(row[col_index(config, "ITEMCOM_QUANTITY")].value))
            add_attribute(c, "unitCode", get_code(str(row[col_index(config, "ITEMCOM_QUANTITY_CODE")].value), unit_codes))

        if not is_cell_empty(str(row[col_index(config, "COMPREL_ITEM_ID")].value)):
            cac = el_tree.SubElement(cac_CatalogueLine, "cac:ComponentRelatedItem")
            add_element(el_tree, cac, "cbc:ID", str(row[col_index(config, "COMPREL_ITEM_ID")].value))
            c = add_element(el_tree, cac, "cbc:Quantity", str(row[col_index(config, "COMPREL_ITEM_QUANTITY")].value))
            add_attribute(c, "unitCode", get_code(str(row[col_index(config, "COMPREL_ITEM_QUANTITY_CODE")].value), unit_codes))

        if not is_cell_empty(str(row[col_index(config, "COMPREL2_ITEM_ID")].value)):
            for value in separated_string(str(row[col_index(config, "COMPREL2_ITEM_ID")].value)):
                cac = el_tree.SubElement(cac_CatalogueLine, "cac:ComponentRelatedItem")
                add_element(el_tree, cac, "cbc:ID", value)

        if not is_cell_empty(str(row[col_index(config, "ASSOCREL_ITEM_ID")].value)):
            for value in separated_string(str(row[col_index(config, "ASSOCREL_ITEM_ID")].value)):
                cac = el_tree.SubElement(cac_CatalogueLine, "cac:AccessoryRelatedItem")
                add_element(el_tree, cac, "cbc:ID", value)

        if not is_cell_empty(str(row[col_index(config, "REQUIREDREL_ITEM_ID")].value)):
            for value in separated_string(str(row[col_index(config, "REQUIREDREL_ITEM_ID")].value)):
                cac = el_tree.SubElement(cac_CatalogueLine, "cac:RequiredRelatedItem")
                add_element(el_tree, cac, "cbc:ID", value)

        if not is_cell_empty(str(row[col_index(config, "REPLACEDREL_ITEM_ID")].value)):
            cac = el_tree.SubElement(cac_CatalogueLine, "cac:ReplacedRelatedItem")
            add_element(el_tree, cac, "cbc:ID", str(row[col_index(config, "REPLACEDREL_ITEM_ID")].value))

        if not is_cell_empty(str(row[col_index(config, "PRICEAMOUNT")].value)):
            add_price(el_tree, cac_CatalogueLine, "", "", str(row[col_index(config, "PRICEAMOUNT")].value), currency_id, str(row[col_index(config, "BASEQUANTITY")].value)
                      , get_code(str(row[col_index(config, "BASEQUANTITY_CODE")].value), unit_codes), get_code(str(row[col_index(config, "PRICETYPE")].value), price_type_codes)
                      , str(row[col_index(config, "PRICE_STARTDATE")].value), str(row[col_index(config, "PRICE_ENDDATE")].value), str(row[col_index(config, "LEADTIMEMEASURE")].value))

        # If more than one price tier
        if not is_cell_empty(str(row[col_index(config, "PRICEAMOUNT_TIER1")].value)):
            add_price(el_tree, cac_CatalogueLine, str(row[col_index(config, "MINIMUMQUANTITY_TIER1")].value), get_code(str(row[col_index(config, "BASEQUANTITY_CODE")].value), unit_codes)
                      , str(row[col_index(config, "PRICEAMOUNT_TIER1")].value), currency_id, str(row[col_index(config, "BASEQUANTITY")].value)
                      , get_code(str(row[col_index(config, "BASEQUANTITY_CODE")].value), unit_codes), get_code(str(row[col_index(config, "PRICETYPE")].value), price_type_codes)
                      , str(row[col_index(config, "PRICE_STARTDATE")].value), str(row[col_index(config, "PRICE_ENDDATE")].value), str(row[col_index(config, "LEADTIMEMEASURE")].value))

        # TIER 2
        if not is_cell_empty(str(row[col_index(config, "PRICEAMOUNT_TIER2")].value)):
            add_price(el_tree, cac_CatalogueLine, str(row[col_index(config, "MINIMUMQUANTITY_TIER2")].value), get_code(str(row[col_index(config, "BASEQUANTITY_CODE")].value), unit_codes)
                      , str(row[col_index(config, "PRICEAMOUNT_TIER2")].value), currency_id, str(row[col_index(config, "BASEQUANTITY")].value)
                      , get_code(str(row[col_index(config, "BASEQUANTITY_CODE")].value), unit_codes), get_code(str(row[col_index(config, "PRICETYPE")].value), price_type_codes)
                      , str(row[col_index(config, "PRICE_STARTDATE")].value), str(row[col_index(config, "PRICE_ENDDATE")].value), str(row[col_index(config, "LEADTIMEMEASURE")].value))

        # TIER 3
        if not is_cell_empty(str(row[col_index(config, "PRICEAMOUNT_TIER3")].value)):
            add_price(el_tree, cac_CatalogueLine, str(row[col_index(config, "MINIMUMQUANTITY_TIER3")].value), get_code(str(row[col_index(config, "BASEQUANTITY_CODE")].value), unit_codes)
                      , str(row[col_index(config, "PRICEAMOUNT_TIER3")].value), currency_id, str(row[col_index(config, "BASEQUANTITY")].value)
                      , get_code(str(row[col_index(config, "BASEQUANTITY_CODE")].value), unit_codes), get_code(str(row[col_index(config, "PRICETYPE")].value), price_type_codes)
                      , str(row[col_index(config, "PRICE_STARTDATE")].value), str(row[col_index(config, "PRICE_ENDDATE")].value), str(row[col_index(config, "LEADTIMEMEASURE")].value))

        #  TIER 4
        if not is_cell_empty(str(row[col_index(config, "PRICEAMOUNT_TIER4")].value)):
            add_price(el_tree, cac_CatalogueLine, str(row[col_index(config, "MINIMUMQUANTITY_TIER4")].value), get_code(str(row[col_index(config, "BASEQUANTITY_CODE")].value), unit_codes)
                      , str(row[col_index(config, "PRICEAMOUNT_TIER4")].value), currency_id, str(row[col_index(config, "BASEQUANTITY")].value)
                      , get_code(str(row[col_index(config, "BASEQUANTITY_CODE")].value), unit_codes), get_code(str(row[col_index(config, "PRICETYPE")].value), price_type_codes)
                      , str(row[col_index(config, "PRICE_STARTDATE")].value), str(row[col_index(config, "PRICE_ENDDATE")].value), str(row[col_index(config, "LEADTIMEMEASURE")].value))

        # Item element
        item = el_tree.SubElement(cac_CatalogueLine, "cac:Item")
        add_element(el_tree, item, "cbc:Description", str(row[col_index(config, "ITEM_DESCRIPTION")].value))
        c = add_element(el_tree, item, "cbc:PackQuantity", str(row[col_index(config, "ITEM_PACKQUANTITY")].value))
        add_attribute(c, "unitCode", get_code(str(row[col_index(config, "ITEM_PACKQUANTITY_CODE")].value), unit_codes))
        add_element(el_tree, item, "cbc:PackSizeNumeric", str(row[col_index(config, "ITEM_PACKSIZENUMERIC")].value))
        add_element(el_tree, item, "cbc:Name", str(row[col_index(config, "ITEM_NAME")].value))
        add_element(el_tree, item, "cbc:Keyword", str(row[col_index(config, "ITEM_KEYWORD")].value))
        add_element(el_tree, item, "cbc:BrandName", str(row[col_index(config, "ITEM_BRANDNAME")].value))

        if not is_cell_empty(str(row[col_index(config, "SELLERSITEMIDENTIFICATION_ID")].value)):
            cac = el_tree.SubElement(item, "cac:SellersItemIdentification")
            add_element(el_tree, cac, "cbc:ID", str(row[col_index(config, "SELLERSITEMIDENTIFICATION_ID")].value))

        if not is_cell_empty(str(row[col_index(config, "MANUFACTURERSITEMIDENTIFICATION_ID")].value)):
            cac = el_tree.SubElement(item, "cac:ManufacturersItemIdentification")
            add_element(el_tree, cac, "cbc:ID", str(row[col_index(config, "MANUFACTURERSITEMIDENTIFICATION_ID")].value))

        if not is_cell_empty(str(row[col_index(config, "STANDARDITEMIDENTIFICATION_ID")].value)):
            cac = el_tree.SubElement(item, "cac:StandardItemIdentification")
            c = add_element(el_tree, cac, "cbc:ID", str(row[col_index(config, "STANDARDITEMIDENTIFICATION_ID")].value))
            add_attribute(c, "schemeID", "0160")  # Only GTIN

        # Product info link
        if not is_cell_empty(str(row[col_index(config, "ITEMSPECIFICATION_EXTERNAL_URI")].value)):
            cac = el_tree.SubElement(item, "cac:ItemSpecificationDocumentReference")
            add_element(el_tree, cac, "cbc:ID", "NA")
            add_element(el_tree, cac, "cbc:DocumentTypeCode", "TRADE_ITEM_DESCRIPTION")
            cac1 = el_tree.SubElement(cac, "cac:Attachment")
            cac2 = el_tree.SubElement(cac1, "cac:ExternalReference")
            add_element(el_tree, cac2, "cbc:URI", str(row[col_index(config, "ITEMSPECIFICATION_EXTERNAL_URI")].value))

        # Product Image link
        if not is_cell_empty(str(row[col_index(config, "ITEMSPECIFICATION_PRODUCT_IMAGE_URI")].value)):
            cac = el_tree.SubElement(item, "cac:ItemSpecificationDocumentReference")
            add_element(el_tree, cac, "cbc:ID", "NA")
            add_element(el_tree, cac, "cbc:DocumentTypeCode", "PRODUCT_IMAGE")
            cac1 = el_tree.SubElement(cac, "cac:Attachment")
            cac2 = el_tree.SubElement(cac1, "cac:ExternalReference")
            add_element(el_tree, cac2, "cbc:URI", str(row[col_index(config, "ITEMSPECIFICATION_PRODUCT_IMAGE_URI")].value))

        # Origin country
        if not is_cell_empty(str(row[col_index(config, "ORIGIN_COUNTRY_CODE")].value)):
            cac = el_tree.SubElement(item, "cac:OriginCountry")
            if len(str(row[col_index(config, "ORIGIN_COUNTRY_CODE")].value)) == 2:
                add_element(el_tree, cac, "cbc:IdentificationCode", str(row[col_index(config, "ORIGIN_COUNTRY_CODE")].value))
            else:
                add_element(el_tree, cac, "cbc:IdentificationCode", get_code(str(row[col_index(config, "ORIGIN_COUNTRY_CODE")].value), country_codes))

        # Varugrupp SSU
        if not is_cell_empty(str(row[col_index(config, "ITEMCLASSIFICATIONCODE_SSU")].value)):
            cac = el_tree.SubElement(item, "cac:CommodityClassification")
            c = add_element(el_tree, cac, "cbc:ItemClassificationCode", str(row[col_index(config, "ITEMCLASSIFICATIONCODE_SSU")].value))
            add_attribute(c, "listID", "SSU")
            add_attribute(c, "name", str(row[col_index(config, "ITEMCLASSIFICATION_VARUGRUPP_DESC")].value))

        # Varugrupp UNCSP
        if not is_cell_empty(str(row[col_index(config, "ITEMCLASSIFICATIONCODE_UNSPSC")].value)):
            cac = el_tree.SubElement(item, "cac:CommodityClassification")
            c = add_element(el_tree, cac, "cbc:ItemClassificationCode", str(row[col_index(config, "ITEMCLASSIFICATIONCODE_UNSPSC")].value))
            add_attribute(c, "listID", "TST")

        # Varugrupp ATC (STL)
        if not is_cell_empty(str(row[col_index(config, "ITEMCLASSIFICATIONCODE_STL")].value)):
            cac = el_tree.SubElement(item, "cac:CommodityClassification")
            c = add_element(el_tree, cac, "cbc:ItemClassificationCode", str(row[col_index(config, "ITEMCLASSIFICATIONCODE_STL")].value))
            add_attribute(c, "listID", "STL")

        # Varugrupp ISO - 9999: 2016 (CC)
        if not is_cell_empty(str(row[col_index(config, "ITEMCLASSIFICATIONCODE_CC")].value)):
            cac = el_tree.SubElement(item, "cac:CommodityClassification")
            c = add_element(el_tree, cac, "cbc:ItemClassificationCode", str(row[col_index(config, "ITEMCLASSIFICATIONCODE_CC")].value))
            add_attribute(c, "listID", "CC")
            add_attribute(c, "listVersionID", "ISO-9999:2016")

        # Contracted item indicator
        if not is_cell_empty(str(row[col_index(config, "CONTRACTED_ITEM")].value)):
            if str(row[col_index(config, "CONTRACTED_ITEM")].value).lower() == "ja":
                cac = el_tree.SubElement(item, "cac:TransactionConditions")
                add_element(el_tree, cac, "cbc:ActionCode", "CT")

        if not is_cell_empty(str(row[col_index(config, "HAZARDOUSITEM_CODE")].value)):
            cac = el_tree.SubElement(item, "cac:HazardousItem")
            add_element(el_tree, cac, "cbc:UNDGCode", str(row[col_index(config, "HAZARDOUSITEM_CODE")].value))
            add_element(el_tree, cac, "cbc:HazardClassID", str(row[col_index(config, "HAZARDOUSITEM_CLASS_ID")].value))

        # VAT category
        if not is_cell_empty(str(row[col_index(config, "CLASSIFIEDTAXCATEGORY_CODE")].value)):
            cac = el_tree.SubElement(item, "cac:ClassifiedTaxCategory")
            add_element(el_tree, cac, "cbc:ID", get_code(str(row[col_index(config, "CLASSIFIEDTAXCATEGORY_CODE")].value), vat_codes))
            add_element(el_tree, cac, "cbc:Percent", str(row[col_index(config, "CLASSIFIEDTAXCATEGORY_CODE")].value))
            cac = el_tree.SubElement(cac, "cac:TaxScheme")
            add_element(el_tree, cac, "cbc:ID", "VAT")

        # SFTI-specific use of additional item Property
        if not is_cell_empty(str(row[col_index(config, "ADD_PROP_VARIABLE_Q")].value)):
            value_string = ""
            if str(row[col_index(config, "ADD_PROP_VARIABLE_Q")].value) == "JA":
                value_string = "true"
            else:
                value_string = "false"
            add_additional_item_prop(el_tree, item, "Variabelmåttvara", "VQ", "GS17009:SFTI", value_string, "SFTI:T0186")

        if not is_cell_empty(str(row[col_index(config, "ADD_PROP_AVAILABILITY")].value)):
            add_additional_item_prop(el_tree, item, str(row[col_index(config, "ADD_PROP_AVAILABILITY")].value),
                                     get_code(str(row[col_index(config, "ADD_PROP_AVAILABILITY")].value), item_availability_codes), "GS14183:SFTI", "true", "SFTI:T0014")

        if not is_cell_empty(str(row[col_index(config, "ADD_PROP_1_TYPE_FROM_TABLE")].value)):
            add_additional_item_prop(el_tree, item, name=str(row[col_index(config, "ADD_PROP_1_TYPE_FROM_TABLE")].value),
                                     name_code=get_code(str(row[col_index(config, "ADD_PROP_1_TYPE_FROM_TABLE")].value), item_attribute_codes),
                                     name_code_list_id=get_code(str(row[col_index(config, "ADD_PROP_1_TYPE_FROM_TABLE")].value), item_attribute_codes, alternate_field="Attr1"),
                                     value=get_code(str(row[col_index(config, "ADD_PROP_1_TYPE_FROM_TABLE")].value), item_attribute_codes, alternate_field="Attr2"),
                                     value_qualifier=get_code(str(row[col_index(config, "ADD_PROP_1_TYPE_FROM_TABLE")].value), item_attribute_codes, alternate_field="Attr3"))

        if not is_cell_empty(str(row[col_index(config, "ADD_PROP_2_TYPE_FROM_TABLE")].value)):
            add_additional_item_prop(el_tree, item, name=str(row[col_index(config, "ADD_PROP_2_TYPE_FROM_TABLE")].value),
                                     name_code=get_code(str(row[col_index(config, "ADD_PROP_2_TYPE_FROM_TABLE")].value), item_attribute_codes),
                                     name_code_list_id=get_code(str(row[col_index(config, "ADD_PROP_2_TYPE_FROM_TABLE")].value), item_attribute_codes, alternate_field="Attr1"),
                                     value=get_code(str(row[col_index(config, "ADD_PROP_2_TYPE_FROM_TABLE")].value), item_attribute_codes, alternate_field="Attr2"),
                                     value_qualifier=get_code(str(row[col_index(config, "ADD_PROP_2_TYPE_FROM_TABLE")].value), item_attribute_codes, alternate_field="Attr3"))

        if not is_cell_empty(str(row[col_index(config, "ADD_PROP_3_TYPE_FROM_TABLE")].value)):
            add_additional_item_prop(el_tree, item, name=str(row[col_index(config, "ADD_PROP_3_TYPE_FROM_TABLE")].value),
                                     name_code=get_code(str(row[col_index(config, "ADD_PROP_3_TYPE_FROM_TABLE")].value), item_attribute_codes),
                                     name_code_list_id=get_code(str(row[col_index(config, "ADD_PROP_3_TYPE_FROM_TABLE")].value), item_attribute_codes, alternate_field="Attr1"),
                                     value=get_code(str(row[col_index(config, "ADD_PROP_3_TYPE_FROM_TABLE")].value), item_attribute_codes, alternate_field="Attr2"),
                                     value_qualifier=get_code(str(row[col_index(config, "ADD_PROP_3_TYPE_FROM_TABLE")].value), item_attribute_codes, alternate_field="Attr3"))

        if not is_cell_empty(str(row[col_index(config, "ADD_PROP_4_TYPE_FROM_TABLE")].value)):
            add_additional_item_prop(el_tree, item, name=str(row[col_index(config, "ADD_PROP_4_TYPE_FROM_TABLE")].value),
                                     name_code=get_code(str(row[col_index(config, "ADD_PROP_4_TYPE_FROM_TABLE")].value), item_attribute_codes),
                                     name_code_list_id=get_code(str(row[col_index(config, "ADD_PROP_4_TYPE_FROM_TABLE")].value), item_attribute_codes, alternate_field="Attr1"),
                                     value=get_code(str(row[col_index(config, "ADD_PROP_4_TYPE_FROM_TABLE")].value), item_attribute_codes, alternate_field="Attr2"),
                                     value_qualifier=get_code(str(row[col_index(config, "ADD_PROP_4_TYPE_FROM_TABLE")].value), item_attribute_codes, alternate_field="Attr3"))

        if not is_cell_empty(str(row[col_index(config, "ADD_PROP_5_TYPE_FROM_TABLE")].value)):
            add_additional_item_prop(el_tree, item, name=str(row[col_index(config, "ADD_PROP_5_TYPE_FROM_TABLE")].value),
                                     name_code=get_code(str(row[col_index(config, "ADD_PROP_5_TYPE_FROM_TABLE")].value), item_attribute_codes),
                                     name_code_list_id=get_code(str(row[col_index(config, "ADD_PROP_5_TYPE_FROM_TABLE")].value), item_attribute_codes, alternate_field="Attr1"),
                                     value=get_code(str(row[col_index(config, "ADD_PROP_5_TYPE_FROM_TABLE")].value), item_attribute_codes, alternate_field="Attr2"),
                                     value_qualifier=get_code(str(row[col_index(config, "ADD_PROP_5_TYPE_FROM_TABLE")].value), item_attribute_codes, alternate_field="Attr3"))

        if not is_cell_empty(str(row[col_index(config, "ADD_PROP_1_NAME")].value)):
            add_additional_item_prop(el_tree, item, str(row[col_index(config, "ADD_PROP_1_NAME")].value), "", "", str(row[col_index(config, "ADD_PROP_1_VALUE")].value),
                                     get_code(str(row[col_index(config, "ADD_PROP_1_NAME")].value), item_property_codes))

        if not is_cell_empty(str(row[col_index(config, "ADD_PROP_2_NAME")].value)):
            add_additional_item_prop(el_tree, item, str(row[col_index(config, "ADD_PROP_2_NAME")].value), "", "", str(row[col_index(config, "ADD_PROP_2_VALUE")].value),
                                     get_code(str(row[col_index(config, "ADD_PROP_2_NAME")].value), item_property_codes))

        if not is_cell_empty(str(row[col_index(config, "ADD_PROP_3_NAME")].value)):
            add_additional_item_prop(el_tree, item, str(row[col_index(config, "ADD_PROP_3_NAME")].value), "", "", str(row[col_index(config, "ADD_PROP_3_VALUE")].value),
                                     get_code(str(row[col_index(config, "ADD_PROP_3_NAME")].value), item_property_codes))

        # Users own text property
        if not is_cell_empty(str(row[col_index(config, "ADD_PROP_1_USERTEXT_NAME")].value)):
            add_additional_item_prop(el_tree, item, str(row[col_index(config, "ADD_PROP_1_USERTEXT_NAME")].value), "", "", str(row[col_index(config, "ADD_PROP_1_USERTEXT_VALUE")].value), "")

        if not is_cell_empty(str(row[col_index(config, "MANUFACTURERPARTY_NAME")].value)):
            cac = el_tree.SubElement(item, "cac:ManufacturerParty")
            cac1 = el_tree.SubElement(cac, "cac:PartyName")
            add_element(el_tree, cac1, "cbc:Name", str(row[col_index(config, "MANUFACTURERPARTY_NAME")].value))

        # CERTIFICATES Environment
        if not is_cell_empty(str(row[col_index(config, "CERTIFICATE_ENV_1")].value)):
            add_item_certificate(el_tree, item, get_code(str(row[col_index(config, "CERTIFICATE_ENV_1")].value), item_certificate_env_codes), "Environmental",
                                 str(row[col_index(config, "CERTIFICATE_ENV_1")].value), "GS1SWEDENT0142")

        if not is_cell_empty(str(row[col_index(config, "CERTIFICATE_ENV_2")].value)):
            add_item_certificate(el_tree, item, get_code(str(row[col_index(config, "CERTIFICATE_ENV_2")].value), item_certificate_env_codes), "Environmental",
                                 str(row[col_index(config, "CERTIFICATE_ENV_2")].value), "GS1SWEDENT0142")

        if not is_cell_empty(str(row[col_index(config, "CERTIFICATE_ENV_3")].value)):
            add_item_certificate(el_tree, item, get_code(str(row[col_index(config, "CERTIFICATE_ENV_3")].value), item_certificate_env_codes), "Environmental",
                                 str(row[col_index(config, "CERTIFICATE_ENV_3")].value), "GS1SWEDENT0142")

        if not is_cell_empty(str(row[col_index(config, "CERTIFICATE_ENV_4")].value)):
            add_item_certificate(el_tree, item, get_code(str(row[col_index(config, "CERTIFICATE_ENV_4")].value), item_certificate_env_codes), "Environmental",
                                 str(row[col_index(config, "CERTIFICATE_ENV_4")].value), "GS1SWEDENT0142")

        if not is_cell_empty(str(row[col_index(config, "CERTIFICATE_ENV_5")].value)):
            add_item_certificate(el_tree, item, get_code(str(row[col_index(config, "CERTIFICATE_ENV_5")].value), item_certificate_env_codes), "Environmental",
                                 str(row[col_index(config, "CERTIFICATE_ENV_5")].value), "GS1SWEDENT0142")

        # Nutrition
        if not is_cell_empty(str(row[col_index(config, "CERTIFICATE_NUTR_1")].value)):
            add_item_certificate(el_tree, item, get_code(str(row[col_index(config, "CERTIFICATE_NUTR_1")].value), item_certificate_nutr_codes), "Nutrition",
                                 str(row[col_index(config, "CERTIFICATE_NUTR_1")].value), "GS1SWEDENT0142")

        if not is_cell_empty(str(row[col_index(config, "CERTIFICATE_NUTR_2")].value)):
            add_item_certificate(el_tree, item, get_code(str(row[col_index(config, "CERTIFICATE_NUTR_2")].value), item_certificate_nutr_codes), "Nutrition",
                                 str(row[col_index(config, "CERTIFICATE_NUTR_2")].value), "GS1SWEDENT0142")

        if not is_cell_empty(str(row[col_index(config, "CERTIFICATE_NUTR_3")].value)):
            add_item_certificate(el_tree, item, get_code(str(row[col_index(config, "CERTIFICATE_NUTR_3")].value), item_certificate_nutr_codes), "Nutrition",
                                 str(row[col_index(config, "CERTIFICATE_NUTR_3")].value), "GS1SWEDENT0142")

        if not is_cell_empty(str(row[col_index(config, "CERTIFICATE_NUTR_4")].value)):
            add_item_certificate(el_tree, item, get_code(str(row[col_index(config, "CERTIFICATE_NUTR_4")].value), item_certificate_nutr_codes), "Nutrition",
                                 str(row[col_index(config, "CERTIFICATE_NUTR_4")].value), "GS1SWEDENT0142")

        if not is_cell_empty(str(row[col_index(config, "CERTIFICATE_NUTR_5")].value)):
            add_item_certificate(el_tree, item, get_code(str(row[col_index(config, "CERTIFICATE_NUTR_5")].value), item_certificate_nutr_codes), "Nutrition",
                                 str(row[col_index(config, "CERTIFICATE_NUTR_5")].value), "GS1SWEDENT0142")

        # Length/Depth
        if not is_cell_empty(str(row[col_index(config, "DIMENSION_ATTR_LN_MEASURE")].value)):
            cac = el_tree.SubElement(item, "cac:Dimension")
            add_element(el_tree, cac, "cbc:AttributeID", "LN")
            c = add_element(el_tree, cac, "cbc:Measure", str(row[col_index(config, "DIMENSION_ATTR_LN_MEASURE")].value))
            add_attribute(c, "unitCode", get_code(str(row[col_index(config, "DIMENSION_ATTR_LN_MEASURE_UOM")].value), unit_codes))

        # Width
        if not is_cell_empty(str(row[col_index(config, "DIMENSION_ATTR_WD_MEASURE")].value)):
            cac = el_tree.SubElement(item, "cac:Dimension")
            add_element(el_tree, cac, "cbc:AttributeID", "WD")
            c = add_element(el_tree, cac, "cbc:Measure", str(row[col_index(config, "DIMENSION_ATTR_WD_MEASURE")].value))
            add_attribute(c, "unitCode", get_code(str(row[col_index(config, "DIMENSION_ATTR_WD_MEASURE_UOM")].value), unit_codes))

        # Height
        if not is_cell_empty(str(row[col_index(config, "DIMENSION_ATTR_HT_MEASURE")].value)):
            cac = el_tree.SubElement(item, "cac:Dimension")
            add_element(el_tree, cac, "cbc:AttributeID", "HT")
            c = add_element(el_tree, cac, "cbc:Measure", str(row[col_index(config, "DIMENSION_ATTR_HT_MEASURE")].value))
            add_attribute(c, "unitCode", get_code(str(row[col_index(config, "DIMENSION_ATTR_HT_MEASURE_UOM")].value), unit_codes))

        # Weight
        if not is_cell_empty(str(row[col_index(config, "DIMENSION_ATTR_GW_MEASURE")].value)):
            cac = el_tree.SubElement(item, "cac:Dimension")
            add_element(el_tree, cac, "cbc:AttributeID", "GW")
            c = add_element(el_tree, cac, "cbc:Measure", str(row[col_index(config, "DIMENSION_ATTR_GW_MEASURE")].value))
            add_attribute(c, "unitCode", get_code(str(row[col_index(config, "DIMENSION_ATTR_GW_MEASURE_UOM")].value), unit_codes))

        # Volume
        if not is_cell_empty(str(row[col_index(config, "DIMENSION_ATTR_ABJ_MEASURE")].value)):
            cac = el_tree.SubElement(item, "cac:Dimension")
            add_element(el_tree, cac, "cbc:AttributeID", "ABJ")
            c = add_element(el_tree, cac, "cbc:Measure", str(row[col_index(config, "DIMENSION_ATTR_ABJ_MEASURE")].value))
            add_attribute(c, "unitCode", get_code(str(row[col_index(config, "DIMENSION_ATTR_ABJ_MEASURE_UOM")].value), unit_codes))

        # net weight
        if not is_cell_empty(str(row[col_index(config, "DIMENSION_ATTR_AAF_MEASURE")].value)):
            cac = el_tree.SubElement(item, "cac:Dimension")
            add_element(el_tree, cac, "cbc:AttributeID", "AAF")
            c = add_element(el_tree, cac, "cbc:Measure", str(row[col_index(config, "DIMENSION_ATTR_AAF_MEASURE")].value))
            add_attribute(c, "unitCode", get_code(str(row[col_index(config, "DIMENSION_ATTR_AAF_MEASURE_UOM")].value), unit_codes))

        # Approx net weight
        if not is_cell_empty(str(row[col_index(config, "DIMENSION_ATTR_APPROX_AAF_MEASURE")].value)):
            cac = el_tree.SubElement(item, "cac:Dimension")
            add_element(el_tree, cac, "cbc:AttributeID", "AAF")
            c = add_element(el_tree, cac, "cbc:Measure", str(row[col_index(config, "DIMENSION_ATTR_APPROX_AAF_MEASURE")].value))
            add_attribute(c, "unitCode", get_code(str(row[col_index(config, "DIMENSION_ATTR_APPROX_AAF_MEASURE_UOM")].value), unit_codes))
            add_element(el_tree, cac, "cbc:Description", "Approximate net weight")

        # net volume
        if not is_cell_empty(str(row[col_index(config, "DIMENSION_ATTR_AAX_MEASURE")].value)):
            cac = el_tree.SubElement(item, "cac:Dimension")
            add_element(el_tree, cac, "cbc:AttributeID", "AAX")
            c = add_element(el_tree, cac, "cbc:Measure", str(row[col_index(config, "DIMENSION_ATTR_AAX_MEASURE")].value))
            add_attribute(c, "unitCode", get_code(str(row[col_index(config, "DIMENSION_ATTR_AAX_MEASURE_UOM")].value), unit_codes))

        # Temperature min max
        if not is_cell_empty(str(row[col_index(config, "DIMENSION_ATTR_TC_MIN_MEASURE")].value)) or not is_cell_empty(str(row[col_index(config, "DIMENSION_ATTR_TC_MAX_MEASURE")].value)):
            cac = el_tree.SubElement(item, "cac:Dimension")
            add_element(el_tree, cac, "cbc:AttributeID", "TC")
            c = add_element(el_tree, cac, "cbc:MinimumMeasure", str(row[col_index(config, "DIMENSION_ATTR_TC_MIN_MEASURE")].value))
            add_attribute(c, "unitCode", "CEL")
            c = add_element(el_tree, cac, "cbc:MaximumMeasure", str(row[col_index(config, "DIMENSION_ATTR_TC_MAX_MEASURE")].value))
            add_attribute(c, "unitCode", "CEL")

        # Humidity min max
        if not is_cell_empty(str(row[col_index(config, "DIMENSION_ATTR_AAO_MIN_MEASURE")].value)) or not is_cell_empty(str(row[col_index(config, "DIMENSION_ATTR_AAO_MAX_MEASURE")].value)):
            cac = el_tree.SubElement(item, "cac:Dimension")
            add_element(el_tree, cac, "cbc:AttributeID", "AAO")
            c = add_element(el_tree, cac, "cbc:MinimumMeasure", str(row[col_index(config, "DIMENSION_ATTR_AAO_MIN_MEASURE")].value))
            add_attribute(c, "unitCode", "P1")
            c = add_element(el_tree, cac, "cbc:MaximumMeasure", str(row[col_index(config, "DIMENSION_ATTR_AAO_MAX_MEASURE")].value))
            add_attribute(c, "unitCode", "P1")

    return el_tree.tostring(root, encoding="utf-8", xml_declaration=True).decode("utf-8")
    #return el_tree.tostring(root, encoding="unicode")


